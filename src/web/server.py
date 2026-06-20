import os
import re
import secrets
import logging
import asyncio
from pathlib import Path
from typing import Dict, Any, Optional, List
import httpx
from fastapi import FastAPI, Depends, HTTPException, status, Body, BackgroundTasks
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel

from src.database.db_manager import DBManager
from src.vk.client_manager import ClientManager

logger = logging.getLogger("toadbot.web.server")

async def send_init_commands(vk_id: int, chat_id: int, client_manager: ClientManager):
    """Асинхронно отправляет инициализационные команды с паузой"""
    try:
        await asyncio.sleep(1.0)
        await client_manager.send_command(vk_id, chat_id, "Моя жаба")
        await asyncio.sleep(1.5)
        await client_manager.send_command(vk_id, chat_id, "Жаба инфо")
    except Exception as e:
        logger.error(f"Ошибка при отправке инициализационных команд для ID {vk_id}: {e}", exc_info=True)

# Модели валидации Pydantic
class ParseTokenSchema(BaseModel):
    url_or_token: str

class AddAccountSchema(BaseModel):
    vk_id: int
    name: str
    token: str
    chat_id: Optional[int] = None
    is_prime: int = 0
    proxy_host: Optional[str] = None
    proxy_port: Optional[int] = None
    proxy_user: Optional[str] = None
    proxy_pass: Optional[str] = None
    proxy_type: Optional[str] = None
    screen_name: Optional[str] = None

class UpdateSettingsSchema(BaseModel):
    auto_feed: Optional[int] = None
    auto_work: Optional[int] = None
    auto_arena: Optional[int] = None
    auto_dungeon: Optional[int] = None
    work_type: Optional[str] = None
    dungeon_type: Optional[str] = None
    arena_league: Optional[str] = None

class GlobalSettingsSchema(BaseModel):
    work_start_grace: int
    work_travel_grace: int
    work_end_grace: int
    min_command_delay: int

class ToggleAccountSchema(BaseModel):
    active: int
    reset_stats: Optional[int] = 0

class ToggleAllBody(BaseModel):
    active: int
    reset_stats: Optional[int] = 0

class TestPhraseSchema(BaseModel):
    vk_id: int
    message: str

class MonitorToggleBody(BaseModel):
    enabled: bool

class MonitoredCommandSchema(BaseModel):
    command: str

class ImportCommandsSchema(BaseModel):
    commands: List[str]

class RecognitionStatusSchema(BaseModel):
    status: str

class ToggleRecognitionSchema(BaseModel):
    in_recognition: int

class TestParseSchema(BaseModel):
    command: str
    text: str

def create_app(db: DBManager, client_manager: ClientManager) -> FastAPI:
    app = FastAPI(title="VK ToadBot Dashboard", version="1.0.0")
    
    @app.on_event("startup")
    async def startup_event():
        logger.info("FastAPI запущен. Инициализация LongPoll-клиентов на event loop Uvicorn...")
        from src.utils.knowledge_base import KnowledgeBase
        await KnowledgeBase.load_from_db(db)
        await client_manager.start_all()
        if hasattr(app.state, "scheduler") and app.state.scheduler:
            app.state.scheduler.start()
            
    @app.on_event("shutdown")
    async def shutdown_event():
        logger.info("FastAPI останавливается. Завершение работы фоновых процессов...")
        if hasattr(app.state, "scheduler") and app.state.scheduler:
            app.state.scheduler.stop()
        await client_manager.stop_all()
    
    @app.middleware("http")
    async def add_no_cache_headers(request, call_next):
        response = await call_next(request)
        if request.url.path.startswith("/static") or request.url.path == "/":
            response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
        return response
    
    # Система базовой авторизации (Basic Auth) для безопасности
    security = HTTPBasic()

    def authenticate(credentials: HTTPBasicCredentials = Depends(security)):
        correct_username = os.getenv("DASHBOARD_USER", "admin")
        correct_password = os.getenv("DASHBOARD_PASS", "admin")
        
        is_correct_username = secrets.compare_digest(credentials.username, correct_username)
        is_correct_password = secrets.compare_digest(credentials.password, correct_password)
        
        if not (is_correct_username and is_correct_password):
            logger.warning(f"Неудачная попытка входа от имени: {credentials.username}")
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Некорректный логин или пароль",
                headers={"WWW-Authenticate": "Basic"},
            )
        return credentials.username

    # Путь к папке со статическими файлами (HTML/CSS/JS)
    static_dir = Path(__file__).resolve().parent / "static"
    static_dir.mkdir(parents=True, exist_ok=True)
    
    # Монтируем статику
    app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

    @app.get("/")
    async def get_index(username: str = Depends(authenticate)):
        """Отдача главной страницы панели управления"""
        index_file = static_dir / "index.html"
        if not index_file.exists():
            # Если файла нет, возвращаем пустую заглушку для первой отрисовки
            return {"status": "error", "message": "index.html not found"}
        return FileResponse(str(index_file))

    # --- REST API ---

    @app.get("/api/accounts")
    async def get_accounts(username: str = Depends(authenticate)) -> List[Dict[str, Any]]:
        """Получение списка всех аккаунтов с их статусами и настройками"""
        try:
            return await db.get_all_accounts()
        except Exception as e:
            logger.error(f"Ошибка при получении списка аккаунтов: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.post("/api/accounts/parse_token")
    async def parse_token_endpoint(data: ParseTokenSchema, username: str = Depends(authenticate)):
        """Парсинг токена/ссылки и запрос к VK API для получения чатов и имени"""
        raw_str = data.url_or_token.strip()
        
        # 1. Извлекаем access_token регулярным выражением
        token = raw_str
        token_match = re.search(r"access_token=([^&]+)", raw_str)
        if token_match:
            token = token_match.group(1)
            
        if not token:
            raise HTTPException(status_code=400, detail="Не удалось извлечь токен доступа из введенных данных.")
            
        # 2. Обращаемся к VK API через httpx
        try:
            async with httpx.AsyncClient() as client:
                # А. Получаем имя, ID и screen_name пользователя
                user_res = await client.get(
                    "https://api.vk.com/method/users.get",
                    params={
                        "access_token": token,
                        "fields": "screen_name",
                        "v": "5.131"
                    }
                )
                user_data = user_res.json()
                
                if "error" in user_data:
                    err_msg = user_data["error"].get("error_msg", "Неизвестная ошибка VK API")
                    logger.warning(f"Ошибка VK API при проверке токена: {err_msg}")
                    raise HTTPException(status_code=400, detail=f"Ошибка VK API: {err_msg}")
                    
                if not user_data.get("response"):
                    raise HTTPException(status_code=400, detail="Неверный ответ от VK API при запросе профиля.")
                    
                user_info = user_data["response"][0]
                vk_id = user_info["id"]
                name = f"{user_info.get('first_name', '')} {user_info.get('last_name', '')}".strip() or f"ID {vk_id}"
                screen_name = user_info.get("screen_name")
                
                # Б. Получаем список бесед пользователя
                conv_res = await client.get(
                    "https://api.vk.com/method/messages.getConversations",
                    params={
                        "access_token": token,
                        "count": 50,
                        "filter": "all",
                        "v": "5.131"
                    }
                )
                conv_data = conv_res.json()
                
                conversations = []
                conversations_error = None
                
                if "error" in conv_data:
                    err_code = conv_data["error"].get("error_code")
                    err_msg = conv_data["error"].get("error_msg", "")
                    logger.warning(f"Не удалось получить список бесед (код {err_code}): {err_msg}")
                    conversations_error = f"Нет доступа к беседам: {err_msg} (Код {err_code}). Убедитесь, что токен имеет права на сообщения (messages)."
                else:
                    items = conv_data.get("response", {}).get("items", [])
                    for item in items:
                        conversation = item.get("conversation", {})
                        peer = conversation.get("peer", {})
                        chat_settings = conversation.get("chat_settings", {})
                        
                        if peer.get("type") == "chat":
                            title = chat_settings.get("title", f"Беседа {peer.get('id')}")
                            conversations.append({
                                "peer_id": peer.get("id"),
                                "title": f"💬 {title}",
                                "type": "chat"
                            })
                        elif peer.get("type") == "user":
                            conversations.append({
                                "peer_id": peer.get("id"),
                                "title": f"👤 Личные сообщения ({peer.get('id')})",
                                "type": "user"
                            })
                            
                return {
                    "status": "success",
                    "vk_id": vk_id,
                    "name": name,
                    "token": token,
                    "screen_name": screen_name,
                    "conversations": conversations,
                    "conversations_error": conversations_error
                }
                
        except httpx.HTTPError as e:
            logger.error(f"Сетевой сбой при обращении к VK API: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Сетевая ошибка при обращении к VK API: {e}")
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Непредвиденная ошибка при разборе токена: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Внутренняя ошибка сервера: {e}")

    @app.post("/api/accounts/add")
    async def add_account(data: AddAccountSchema, username: str = Depends(authenticate)):
        """Добавление нового аккаунта и его мгновенный запуск"""
        try:
            # 1. Записываем в базу данных
            await db.add_account(
                vk_id=data.vk_id,
                name=data.name,
                token=data.token,
                chat_id=data.chat_id,
                is_prime=data.is_prime,
                proxy_host=data.proxy_host,
                proxy_port=data.proxy_port,
                proxy_user=data.proxy_user,
                proxy_pass=data.proxy_pass,
                proxy_type=data.proxy_type,
                screen_name=data.screen_name
            )
            
            # 2. Читаем полную запись из БД (со всеми настройками автоматизации по умолчанию)
            account_data = await db.get_account(data.vk_id)
            if not account_data:
                raise HTTPException(status_code=500, detail="Ошибка при чтении созданной записи")
            
            # 3. Мгновенно запускаем фоновый LongPoll в ClientManager
            success = await client_manager.start_account(account_data)
            
            return {
                "status": "success",
                "message": f"Аккаунт {data.name} успешно добавлен.",
                "bot_started": success
            }
        except Exception as e:
            logger.error(f"Ошибка при добавлении аккаунта: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Ошибка сервера: {e}")
            # 2. Читаем полную запись из БД (со всеми настройками автоматизации по умолчанию)
            account_data = await db.get_account(data.vk_id)
            if not account_data:
                raise HTTPException(status_code=500, detail="Ошибка при чтении созданной записи")
            
            # 3. Мгновенно запускаем фоновый LongPoll в ClientManager
            success = await client_manager.start_account(account_data)
            
            return {
                "status": "success",
                "message": f"Аккаунт {data.name} успешно добавлен.",
                "bot_started": success
            }
        except Exception as e:
            logger.error(f"Ошибка при добавлении аккаунта: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Ошибка сервера: {e}")

    @app.get("/api/global/settings")
    async def get_global_settings_endpoint(username: str = Depends(authenticate)) -> Dict[str, int]:
        """Получение глобальных настроек форы"""
        try:
            return await db.get_global_settings()
        except Exception as e:
            logger.error(f"Ошибка при получении глобальных настроек: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.post("/api/global/settings")
    async def update_global_settings_endpoint(
        data: GlobalSettingsSchema,
        username: str = Depends(authenticate)
    ):
        """Обновление глобальных настроек форы"""
        try:
            updates = data.model_dump()
            await db.update_global_settings(updates)
            await db.log_action(0, "settings", f"Изменены глобальные настройки форы: {updates}")
            return {"status": "success", "message": "Глобальные настройки успешно сохранены"}
        except Exception as e:
            logger.error(f"Ошибка при сохранении глобальных настроек: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка сервера")

    @app.post("/api/accounts/{vk_id}/settings")
    async def update_account_settings(
        vk_id: int, 
        settings_data: UpdateSettingsSchema, 
        username: str = Depends(authenticate)
    ):
        """Обновление настроек автоматизации аккаунта"""
        if vk_id <= 0:
            return {"status": "ignored", "message": "Для общего аккаунта используются глобальные настройки"}
        try:
            # Превращаем Pydantic схему в словарь, исключая неуказанные значения
            updates = settings_data.model_dump(exclude_unset=True)
            if not updates:
                return {"status": "ignored", "message": "Нет данных для обновления"}
                
            await db.update_settings(vk_id, updates)
            await db.log_action(vk_id, "settings", f"Изменены параметры автоматизации: {updates}")
            return {"status": "success", "message": "Настройки успешно сохранены"}
        except Exception as e:
            logger.error(f"Ошибка при обновлении настроек ID {vk_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Ошибка сервера: {e}")

    @app.post("/api/accounts/{vk_id}/toggle")
    async def toggle_account(vk_id: int, data: ToggleAccountSchema, username: str = Depends(authenticate)):
        """Включение или выключение активности аккаунта (запуск/остановка бота)"""
        if vk_id <= 0:
            raise HTTPException(status_code=400, detail="Невозможно изменить статус виртуального или системного аккаунта.")
        try:
            active = data.active
            reset_stats = data.reset_stats
            
            # 1. Меняем статус активности в БД
            await db.toggle_account_active(vk_id, active)
            
            # 2. Сбрасываем статы при включении и согласии
            if active == 1 and reset_stats == 1:
                await db.reset_account_stats(vk_id)
            
            # 3. Запускаем или останавливаем бота
            if active == 1:
                account_data = await db.get_account(vk_id)
                if not account_data:
                    raise HTTPException(status_code=404, detail="Аккаунт не найден")
                success = await client_manager.start_account(account_data)
                msg = "Бот успешно запущен." if success else "Ошибка запуска бота."
                
                # Инициализация команд при сбросе статистики
                if success and reset_stats == 1 and account_data.get("chat_id"):
                    asyncio.create_task(send_init_commands(vk_id, account_data["chat_id"], client_manager))
            else:
                success = await client_manager.stop_account(vk_id)
                msg = "Бот остановлен." if success else "Бот уже был остановлен."
                
            return {
                "status": "success" if success else "warning",
                "message": msg,
                "active": active
            }
        except Exception as e:
            logger.error(f"Ошибка при переключении статуса ID {vk_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Ошибка сервера: {e}")

    @app.post("/api/accounts/toggle_all")
    async def toggle_all_accounts(body: ToggleAllBody, username: str = Depends(authenticate)):
        """Массовый запуск или остановка всех аккаунтов"""
        try:
            active = body.active
            reset_stats = body.reset_stats
            
            accounts = await db.get_all_accounts()
            if not accounts:
                return {"status": "success", "message": "Нет подключенных аккаунтов."}
                
            success_count = 0
            for acc in accounts:
                vk_id = acc["vk_id"]
                
                await db.toggle_account_active(vk_id, active)
                
                if active == 1:
                    if reset_stats == 1:
                        await db.reset_account_stats(vk_id)
                    acc_data = await db.get_account(vk_id)
                    success = await client_manager.start_account(acc_data)
                    if success:
                        success_count += 1
                        # Инициализация команд при сбросе статистики
                        if reset_stats == 1 and acc_data.get("chat_id"):
                            asyncio.create_task(send_init_commands(vk_id, acc_data["chat_id"], client_manager))
                else:
                    success = await client_manager.stop_account(vk_id)
                    if success:
                        success_count += 1
                        
            action_word = "запущены" if active == 1 else "остановлены"
            reset_word = " со сбросом статистики" if (active == 1 and reset_stats == 1) else ""
            msg = f"Массовое действие: все боты успешно {action_word}{reset_word}!"
            
            await db.log_action(0, "system", f"⚠️ Массовое действие: все боты {action_word}{reset_word}.")
            
            return {
                "status": "success",
                "message": msg,
                "active": active
            }
        except Exception as e:
            logger.error(f"Ошибка при массовом переключении аккаунтов: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка сервера при массовом переключении")

    @app.post("/api/accounts/{vk_id}/delete")
    async def delete_account(vk_id: int, username: str = Depends(authenticate)):
        """Полное удаление аккаунта из системы"""
        if vk_id <= 0:
            raise HTTPException(status_code=400, detail="Невозможно удалить общий или системный аккаунт.")
        try:
            # 1. Сначала останавливаем процесс LongPoll, если он запущен
            await client_manager.stop_account(vk_id)
            
            # 2. Удаляем запись из БД
            await db.delete_account(vk_id)
            return {"status": "success", "message": "Аккаунт успешно удален из системы."}
        except Exception as e:
            logger.error(f"Ошибка при удалении аккаунта ID {vk_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Ошибка сервера: {e}")

    @app.get("/api/logs")
    async def get_logs(limit: int = 50, username: str = Depends(authenticate)):
        """Получение последних логов системы"""
        try:
            return await db.get_logs(limit=limit)
        except Exception as e:
            logger.error(f"Ошибка при получении логов: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка сервера при чтении логов")
            
    @app.post("/api/logs/clear")
    async def clear_logs(vk_id: Optional[int] = None, username: str = Depends(authenticate)):
        """Полная и безвозвратная очистка логов в БД (для конкретного аккаунта или всех)"""
        try:
            if vk_id and vk_id > 0:
                await db.clear_logs(vk_id)
                await db.log_action(vk_id, "system", "История логов этого аккаунта безвозвратно очищена.")
                return {"status": "success", "message": "🗑️ История логов этого аккаунта успешно очищена!"}
            else:
                await db.clear_all_logs()
                await db.log_action(0, "system", "Вся системная история логов безвозвратно очищена.")
                return {"status": "success", "message": "🗑️ Все системные логи успешно очищены!"}
        except Exception as e:
            logger.error(f"Ошибка при очистке логов: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка сервера при очистке логов")

    @app.post("/api/debug/monitor/clear")
    async def clear_monitor(username: str = Depends(authenticate)):
        """Очистка базы данных мониторинга и удаление файла отчета Excel"""
        try:
            from pathlib import Path
            await db.clear_monitored_responses()
            report_path = Path("monitor_report.xlsx")
            if report_path.exists():
                report_path.unlink()
            return {"status": "success", "message": "База данных и отчет мониторинга успешно очищены!"}
        except Exception as e:
            logger.error(f"Ошибка при очистке мониторинга: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка сервера")

    # --- API Эндпоинты отслеживания команд (Новый мониторинг) ---

    @app.get("/api/monitor/commands")
    async def get_monitored_commands(username: str = Depends(authenticate)):
        """Получение списка всех отслеживаемых команд и их вариаций"""
        try:
            return await db.get_monitored_commands()
        except Exception as e:
            logger.error(f"Ошибка получения отслеживаемых команд: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.post("/api/monitor/commands")
    async def add_monitored_command(payload: MonitoredCommandSchema, username: str = Depends(authenticate)):
        """Добавление новой команды в список отслеживаемых"""
        cmd = payload.command.strip()
        if not cmd:
            raise HTTPException(status_code=400, detail="Команда не может быть пустой")
        try:
            await db.add_monitored_command(cmd)
            return {"status": "success", "message": f"Команда '{cmd}' добавлена"}
        except Exception as e:
            logger.error(f"Ошибка при добавлении команды: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.delete("/api/monitor/commands/{command_id}")
    async def delete_monitored_command(command_id: int, username: str = Depends(authenticate)):
        """Удаление команды из отслеживаемых по ID"""
        try:
            await db.delete_monitored_command(command_id)
            return {"status": "success", "message": "Команда успешно удалена"}
        except Exception as e:
            logger.error(f"Ошибка при удалении команды ID {command_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.get("/api/monitor/status")
    async def get_monitor_status(username: str = Depends(authenticate)):
        """Получить текущий статус Монитора"""
        enabled = await db.is_monitor_mode_enabled()
        return {"enabled": enabled}

    @app.post("/api/monitor/toggle")
    async def toggle_monitor_mode(body: MonitorToggleBody, username: str = Depends(authenticate)):
        """Включить или выключить режим мониторинга"""
        await db.set_monitor_mode_enabled(body.enabled)
        return {"enabled": body.enabled}

    @app.delete("/api/monitor/responses/{response_id}")
    async def delete_monitored_response(response_id: int, username: str = Depends(authenticate)):
        """Удаление конкретного варианта ответа по ID"""
        try:
            await db.delete_monitored_response(response_id)
            return {"status": "success", "message": "Вариант ответа успешно удален"}
        except Exception as e:
            logger.error(f"Ошибка при удалении варианта ответа ID {response_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.post("/api/monitor/responses/{response_id}/recognition")
    async def update_monitored_response_recognition(response_id: int, payload: RecognitionStatusSchema, username: str = Depends(authenticate)):
        """Обновление статуса распознавания конкретного варианта ответа"""
        try:
            await db.update_monitored_response_recognition(response_id, payload.status)
            return {"status": "success", "message": f"Статус изменен на '{payload.status}'"}
        except Exception as e:
            logger.error(f"Ошибка обновления статуса распознавания варианта ответа ID {response_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.get("/api/monitor/debug/unrecognized")
    async def get_unrecognized_monitor_variations_endpoint(username: str = Depends(authenticate)):
        """Получение списка команд и вариаций ответов, которые не удалось распознать по правилам (статус 'Нет')"""
        try:
            return await db.get_unrecognized_monitor_variations()
        except Exception as e:
            logger.error(f"Ошибка получения нераспознанных вариаций для отладки: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.post("/api/monitor/commands/{command_id}/recognition")
    async def update_monitored_command_recognition(command_id: int, payload: RecognitionStatusSchema, username: str = Depends(authenticate)):
        """Обновление статуса распознавания отслеживаемой команды"""
        try:
            await db.update_monitored_command_recognition(command_id, payload.status)
            return {"status": "success", "message": f"Статус изменен на '{payload.status}'"}
        except Exception as e:
            logger.error(f"Ошибка обновления статуса распознавания команды ID {command_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.post("/api/monitor/commands/{command_id}/recognition/toggle")
    async def toggle_monitored_command_in_recognition(command_id: int, payload: ToggleRecognitionSchema, username: str = Depends(authenticate)):
        """Добавление/удаление команды из вкладки Распознавание"""
        try:
            await db.toggle_monitored_command_in_recognition(command_id, payload.in_recognition)
            return {"status": "success", "message": f"Статус in_recognition изменен на {payload.in_recognition}"}
        except Exception as e:
            logger.error(f"Ошибка при переключении in_recognition для команды ID {command_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.post("/api/monitor/commands/{command_id}/description")
    async def update_command_description(command_id: int, payload: dict, username: str = Depends(authenticate)):
        """Обновление описания команды"""
        try:
            description = str(payload.get("description", ""))[:500]
            async with db._connect() as conn:
                await conn.execute(
                    "UPDATE monitored_commands SET description = ? WHERE id = ?",
                    (description, command_id)
                )
                await conn.commit()
            return {"status": "success", "message": "Описание обновлено"}
        except Exception as e:
            logger.error(f"Ошибка обновления описания команды ID {command_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.get("/api/monitor/commands/{command_id}/recognition/rules")
    async def get_monitored_command_recognition_rules(command_id: int, username: str = Depends(authenticate)):
        """Получение правил распознавания для команды"""
        try:
            rules = await db.get_recognition_rules_for_command(command_id)
            return rules
        except Exception as e:
            logger.error(f"Ошибка получения правил распознавания команды ID {command_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    # ── API БД «Выпадающий лут» ──

    @app.get("/api/loot/groups")
    async def get_loot_groups_route(username: str = Depends(authenticate)):
        """Список всех групп лута с вложенными предметами"""
        try:
            return await db.get_loot_groups()
        except Exception as e:
            logger.error(f"Ошибка получения групп лута: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.post("/api/loot/groups")
    async def create_loot_group_route(payload: dict, username: str = Depends(authenticate)):
        """Создать новую группу лута"""
        try:
            group_key = (payload.get("group_key") or "").strip()
            name = (payload.get("name") or "").strip()
            pattern_regex = payload.get("pattern_regex") or ""
            description = payload.get("description") or ""
            if not group_key or not name:
                raise HTTPException(status_code=400, detail="group_key и name обязательны")
            new_id = await db.create_loot_group(group_key, name, pattern_regex, description)
            return {"success": True, "id": new_id}
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Ошибка создания группы лута: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.put("/api/loot/groups/{group_id}")
    async def update_loot_group_route(group_id: int, payload: dict, username: str = Depends(authenticate)):
        """Обновить группу лута"""
        try:
            await db.update_loot_group(
                group_id,
                name=payload.get("name"),
                pattern_regex=payload.get("pattern_regex"),
                description=payload.get("description"),
            )
            return {"success": True}
        except Exception as e:
            logger.error(f"Ошибка обновления группы лута {group_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.delete("/api/loot/groups/{group_id}")
    async def delete_loot_group_route(group_id: int, username: str = Depends(authenticate)):
        """Удалить группу лута (CASCADE → удалятся и предметы)"""
        try:
            await db.delete_loot_group(group_id)
            return {"success": True}
        except Exception as e:
            logger.error(f"Ошибка удаления группы лута {group_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.post("/api/loot/items")
    async def create_loot_item_route(payload: dict, username: str = Depends(authenticate)):
        """Создать предмет лута"""
        try:
            group_id = payload.get("group_id")
            emoji = (payload.get("emoji") or "").strip()
            db_column = (payload.get("db_column") or "").strip()
            item_name = payload.get("item_name") or ""
            item_regex = payload.get("item_regex") or ""
            if not group_id or not emoji or not db_column:
                raise HTTPException(status_code=400, detail="group_id, emoji, db_column обязательны")
            new_id = await db.create_loot_item(group_id, emoji, db_column, item_name, item_regex)
            return {"success": True, "id": new_id}
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Ошибка создания предмета лута: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.put("/api/loot/items/{item_id}")
    async def update_loot_item_route(item_id: int, payload: dict, username: str = Depends(authenticate)):
        """Обновить предмет лута"""
        try:
            await db.update_loot_item(
                item_id,
                emoji=payload.get("emoji"),
                db_column=payload.get("db_column"),
                item_name=payload.get("item_name"),
                item_regex=payload.get("item_regex"),
            )
            return {"success": True}
        except Exception as e:
            logger.error(f"Ошибка обновления предмета лута {item_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.delete("/api/loot/items/{item_id}")
    async def delete_loot_item_route(item_id: int, username: str = Depends(authenticate)):
        """Удалить предмет лута"""
        try:
            await db.delete_loot_item(item_id)
            return {"success": True}
        except Exception as e:
            logger.error(f"Ошибка удаления предмета лута {item_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    # ── API связки «команда → группы лута» (command_loot_config) ──

    @app.get("/api/monitor/commands/{command_id}/loot-config")
    async def get_command_loot_config_route(command_id: int, username: str = Depends(authenticate)):
        """Получить группы лута, привязанные к команде"""
        try:
            return await db.get_command_loot_config(command_id)
        except Exception as e:
            logger.error(f"Ошибка получения loot-config для команды {command_id}: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.post("/api/monitor/commands/{command_id}/loot-config")
    async def add_command_loot_config_route(command_id: int, payload: dict, username: str = Depends(authenticate)):
        """Привязать группу лута к команде"""
        try:
            group_id = payload.get("group_id")
            section_header = payload.get("section_header") or "Ты получил"
            recipient_mode = payload.get("recipient_mode") or "pending"
            if not group_id:
                raise HTTPException(status_code=400, detail="group_id обязателен")
            ok = await db.add_command_loot_config(command_id, group_id, section_header, recipient_mode)
            if not ok:
                raise HTTPException(status_code=409, detail="Эта группа уже привязана к команде")
            return {"success": True}
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Ошибка привязки loot-config: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.put("/api/monitor/commands/{command_id}/loot-config/{group_id}")
    async def update_command_loot_config_route(command_id: int, group_id: int, payload: dict, username: str = Depends(authenticate)):
        """Обновить настройки (section_header, recipient_mode) привязки группы лута к команде"""
        try:
            section_header = payload.get("section_header")
            recipient_mode = payload.get("recipient_mode")
            # Находим config_id по паре (command_id, group_id)
            configs = await db.get_command_loot_config(command_id)
            config_id = None
            for c in configs:
                if c.get("group_id") == group_id:
                    config_id = c.get("id")
                    break
            if config_id is None:
                raise HTTPException(status_code=404, detail="Привязка не найдена")
            await db.update_command_loot_config(config_id, section_header=section_header, recipient_mode=recipient_mode)
            return {"success": True}
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Ошибка обновления loot-config: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.delete("/api/monitor/commands/{command_id}/loot-config/{group_id}")
    async def remove_command_loot_config_route(command_id: int, group_id: int, username: str = Depends(authenticate)):
        """Отвязать группу лута от команды"""
        try:
            await db.remove_command_loot_config(command_id, group_id)
            return {"success": True}
        except Exception as e:
            logger.error(f"Ошибка отвязки loot-config: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.post("/api/monitor/test-parse")
    async def test_parse_endpoint(payload: TestParseSchema, username: str = Depends(authenticate)):
        """Тестирование распознавания на переданном тексте"""
        try:
            cmd_name = payload.command.strip()
            text = payload.text.strip()
            
            if not cmd_name:
                return {"success": False, "error": "Имя команды не может быть пустым."}
            if not text:
                return {"success": False, "error": "Текст ответа не может быть пустым."}
                
            # 1. Ищем команду в БД
            async with db._connect() as conn:
                async with conn.execute(
                    "SELECT id, in_recognition, command FROM monitored_commands WHERE command = ? COLLATE NOCASE", (cmd_name,)
                ) as cursor:
                    cmd_row = await cursor.fetchone()
                    
                if not cmd_row:
                    return {"success": False, "error": f"Команда '{cmd_name}' отсутствует в списке распознаваемых."}
                    
                cmd_id = cmd_row["id"]
                in_recognition = cmd_row["in_recognition"]
                cmd_real_name = cmd_row["command"]
                
                if in_recognition != 1:
                    return {"success": False, "error": f"Для команды '{cmd_real_name}' выключено распознавание."}
                    
                # 2. Получаем все подразделы и правила
                async with conn.execute("""
                    SELECT s.id as subcommand_id, s.name as subcommand_name, r.id as rule_id, r.pattern, r.variable_name, r.output_value
                    FROM recognition_subcommands s
                    LEFT JOIN recognition_rules r ON s.id = r.subcommand_id
                    WHERE s.command_id = ?
                """, (cmd_id,)) as cursor:
                    rows = [dict(row) for row in await cursor.fetchall()]
                    
            # Группируем правила по подразделам
            subcommands_map = {}
            for r in rows:
                sub_name = r["subcommand_name"]
                if sub_name not in subcommands_map:
                    subcommands_map[sub_name] = {
                        "subcommand_name": sub_name,
                        "rules": []
                    }
                if r["rule_id"] is not None:
                    subcommands_map[sub_name]["rules"].append({
                        "rule_id": r["rule_id"],
                        "pattern": r["pattern"],
                        "variable_name": r["variable_name"],
                        "output_value": r["output_value"]
                    })
                    
            import re
            
            # Проверяем совпадение правил
            subcommands_result = []
            any_matched = False
            
            for sub_name, sub_data in subcommands_map.items():
                matched_rule = None
                # Пробуем найти совпадающее правило
                for rule in sub_data["rules"]:
                    try:
                        text_clean = text.replace("\r", "")
                        pattern = rule["pattern"].strip()
                        if not pattern.endswith("$") and not pattern.endswith(".*"):
                            pattern = pattern + "$"
                        regex = re.compile(pattern, re.IGNORECASE | re.MULTILINE)
                        match = regex.search(text_clean)
                        if match:
                            groups = match.groupdict()
                            # Форматируем output_value
                            formatted_output = rule["output_value"]
                            try:
                                formatted_output = rule["output_value"].format(**groups)
                            except Exception:
                                pass
                                
                            matched_rule = {
                                "rule_id": rule["rule_id"],
                                "pattern": rule["pattern"],
                                "variable_name": rule["variable_name"],
                                "output_value": formatted_output,
                                "captured_groups": groups
                            }
                            any_matched = True
                            break
                    except Exception:
                        pass
                    
                subcommands_result.append({
                    "subcommand_name": sub_name,
                    "matched": matched_rule is not None,
                    "matched_rule": matched_rule
                })
                
            # Вычисляем итоговый статус
            recognized = False
            if cmd_real_name == "Жаба инфо":
                from src.utils.toad_info_parser import parse_toad_info
                parsed_data = parse_toad_info(text)
                recognized = parsed_data is not None
            elif cmd_real_name == "Моя жаба":
                from src.utils.toad_info_parser import parse_my_toad
                parsed_data = parse_my_toad(text)
                recognized = parsed_data is not None
            elif cmd_real_name == "Мой инвентарь":
                from src.utils.toad_info_parser import parse_inventory
                parsed_data = parse_inventory(text)
                recognized = parsed_data is not None
            elif cmd_real_name == "Мое снаряжение":
                from src.utils.toad_info_parser import parse_equipment
                parsed_data = parse_equipment(text)
                recognized = parsed_data is not None
            elif cmd_real_name == "Покормить жабу":
                from src.utils.toad_info_parser import parse_feed
                parsed_data = parse_feed(text)
                recognized = parsed_data is not None
            elif cmd_real_name == "Дейлики":
                from src.utils.toad_info_parser import parse_dailies
                parsed_data = parse_dailies(text)
                recognized = parsed_data is not None
            elif cmd_real_name == "Моя семья":
                from src.utils.toad_info_parser import parse_family
                parsed_data = parse_family(text, "")
                recognized = parsed_data is not None
            elif cmd_real_name == "Моя банда":
                from src.utils.toad_info_parser import parse_gang
                parsed_data = parse_gang(text)
                recognized = parsed_data is not None
            else:
                recognized = any_matched
                
            return {
                "success": True,
                "command_name": cmd_real_name,
                "recognized": recognized,
                "subcommands": subcommands_result
            }
            
        except Exception as e:
            logger.error(f"Ошибка при интерактивном тестировании распознавания: {e}", exc_info=True)
            return {"success": False, "error": f"Внутренняя ошибка сервера: {str(e)}"}

    @app.get("/api/monitor/delta-audit")
    async def get_delta_audit(limit: int = 200, username: str = Depends(authenticate)):
        """Возвращает записи из delta_audit (аудит инкрементов и ненаходов)"""
        try:
            async with db._connect_rec() as conn:
                async with conn.execute(
                    """SELECT id, vk_id, command, event_type, field_name,
                              delta_value, old_value, new_value, raw_text, created_at
                       FROM delta_audit
                       ORDER BY id DESC LIMIT ?""",
                    (min(limit, 500),)
                ) as cursor:
                    rows = [dict(r) for r in await cursor.fetchall()]
            return {"success": True, "rows": rows}
        except Exception as e:
            logger.error(f"Ошибка получения delta_audit: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.delete("/api/monitor/delta-audit")
    async def clear_delta_audit(username: str = Depends(authenticate)):
        """Очищает таблицу delta_audit"""
        try:
            async with db._connect_rec() as conn:
                await conn.execute("DELETE FROM delta_audit")
                await conn.commit()
            return {"success": True}
        except Exception as e:
            logger.error(f"Ошибка очистки delta_audit: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.get("/api/monitor/download")
    async def download_monitor_report(username: str = Depends(authenticate)):
        """Скачать Excel отчет монитора (генерируется на лету)"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from fastapi.responses import FileResponse
        
        def sanitize_sheet_name(name: str) -> str:
            # remove invalid chars: \ / ? * : [ ]
            for c in r"\/?:*[]":
                name = name.replace(c, "")
            name = name.strip()[:30] # Excel limit is 31
            if not name:
                name = "Command"
            return name

        try:
            commands = await db.get_monitored_commands()
            if not commands:
                raise HTTPException(status_code=400, detail="Нет данных для скачивания. Добавьте отслеживаемые команды.")
            
            wb = openpyxl.Workbook()
            # Remove default sheet
            if wb.active:
                wb.remove(wb.active)
                
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2a1b3d", end_color="2a1b3d", fill_type="solid") # Темный фиолетовый под тему
            
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            for cmd_entry in commands:
                cmd_text = cmd_entry["command"]
                sheet_name = sanitize_sheet_name(cmd_text)
                
                # Защита от дубликатов имен листов
                orig_name = sheet_name
                counter = 1
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{orig_name[:27]}_{counter}"
                    counter += 1
                    
                ws = wb.create_sheet(title=sheet_name)
                ws.views.sheetView[0].showGridLines = True
                
                # 1. Извлекаем историю для каждой вариации и находим максимальное количество текстов (редакций)
                variations_data = []
                max_texts = 1
                for var in cmd_entry["variations"]:
                    edits = var.get("response_history", [var["response_text"]])
                    if len(edits) > max_texts:
                        max_texts = len(edits)
                    variations_data.append((
                        edits,
                        var["match_count"],
                        var.get("last_mention_at") or "",
                        var.get("recognition_status") or "Не распознаем"
                    ))
                
                # Сортируем по убыванию частоты совпадений
                variations_data.sort(key=lambda x: x[1], reverse=True)
                
                total_cols = max_texts + 4
                
                # 2. Оформление шапки метаданных (строка 1)
                ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
                cell_title = ws.cell(row=1, column=1, value=f"Команда: {cmd_text}")
                cell_title.font = Font(name="Calibri", size=12, bold=True, color="2a1b3d")
                cell_title.alignment = Alignment(horizontal='left', vertical='center', indent=1)
                ws.row_dimensions[1].height = 24
                
                # Закрашиваем объединенные ячейки метаданных
                for col_idx in range(1, total_cols + 1):
                    ws.cell(row=1, column=col_idx).fill = PatternFill(start_color="e2d9f3", end_color="e2d9f3", fill_type="solid")
                    ws.cell(row=1, column=col_idx).border = thin_border
                
                ws.row_dimensions[2].height = 10  # Разделительный отступ
                
                # 3. Формируем заголовки динамически на 3 строке
                headers = ["Вариант", "Текст"]
                for i in range(1, max_texts):
                    headers.append(f"Текст{i}")
                headers.append("Дата")
                headers.append("Совпадений")
                headers.append("Распознавание")
                
                ws.row_dimensions[3].height = 28
                for col_idx, header in enumerate(headers, 1):
                     cell = ws.cell(row=3, column=col_idx, value=header)
                     cell.font = header_font
                     cell.fill = header_fill
                     cell.alignment = center_align
                     cell.border = thin_border
                     
                # 4. Добавляем строки с автозаполнением пустотами начиная с 4 строки
                for idx, (edits, match_count, last_mention, status) in enumerate(variations_data, 1):
                     row_idx = 3 + idx
                     ws.row_dimensions[row_idx].height = 22
                     
                     # Вариант (Колонка 1)
                     cell_var = ws.cell(row=row_idx, column=1, value=idx)
                     cell_var.font = Font(name="Calibri", size=10)
                     cell_var.alignment = center_align
                     cell_var.border = thin_border
                     
                     # Тексты (Колонки 2 до max_texts + 1)
                     for i in range(max_texts):
                          text_val = edits[i] if i < len(edits) else ""
                          cell_text = ws.cell(row=row_idx, column=2 + i, value=text_val)
                          cell_text.font = Font(name="Calibri", size=10)
                          cell_text.alignment = left_align
                          cell_text.border = thin_border
                          
                     # Дата (Колонка max_texts + 2)
                     cell_date = ws.cell(row=row_idx, column=max_texts + 2, value=last_mention)
                     cell_date.font = Font(name="Calibri", size=10)
                     cell_date.alignment = center_align
                     cell_date.border = thin_border
                     
                     # Совпадений (Колонка max_texts + 3)
                     cell_count = ws.cell(row=row_idx, column=max_texts + 3, value=match_count)
                     cell_count.font = Font(name="Calibri", size=10)
                     cell_count.alignment = center_align
                     cell_count.border = thin_border

                     # Распознавание (Колонка max_texts + 4)
                     cell_status = ws.cell(row=row_idx, column=max_texts + 4, value=status)
                     cell_status.font = Font(name="Calibri", size=10)
                     cell_status.alignment = center_align
                     cell_status.border = thin_border
                     
                for col in ws.columns:
                     max_len = 0
                     col_letter = get_column_letter(col[0].column)
                     # Пропускаем первые 2 строки (метаданные), чтобы не раздувать ширину колонок
                     for cell in col[2:]:
                          val_str = str(cell.value or '')
                          lines = val_str.split('\n')
                          val_len = max(len(l) for l in lines) if lines else 0
                          if val_len > max_len:
                               max_len = val_len
                     col_width = min(max(max_len + 3, 10), 60)
                     ws.column_dimensions[col_letter].width = col_width
                     
            import io
            from fastapi.responses import StreamingResponse
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=monitor_report.xlsx"}
            )
        except HTTPException as he:
            raise he
        except Exception as e:
            logger.error(f"Ошибка генерации отчета Excel: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка при генерации Excel")

    @app.post("/api/monitor/commands/import")
    async def import_monitored_commands(payload: ImportCommandsSchema, username: str = Depends(authenticate)):
        """Массовый импорт команд для отслеживания"""
        if not payload.commands:
            raise HTTPException(status_code=400, detail="Список команд пуст")
        try:
            added_count = await db.add_monitored_commands_batch(payload.commands)
            return {"status": "success", "message": f"Добавлено новых команд: {added_count}"}
        except Exception as e:
            logger.error(f"Ошибка при массовом импорте команд: {e}", exc_info=True)
            raise HTTPException(status_code=500, detail="Ошибка базы данных")

    @app.post("/api/monitor/reset")
    async def reset_monitor_mode(username: str = Depends(authenticate)):
        """Сбросить сохраненные вариации ответов монитора"""
        try:
            await db.clear_monitored_responses()
            report_path = Path("monitor_report.xlsx")
            if report_path.exists():
                report_path.unlink()
            return {"status": "success", "message": "Статистика ответов сброшена!"}
        except Exception as e:
            logger.error(f"Ошибка при сбросе статистики монитора: {e}")
            raise HTTPException(status_code=500, detail="Ошибка при сбросе статистики")

    @app.post("/api/accounts/{vk_id}/like_group")
    async def like_group_endpoint(
        vk_id: int,
        background_tasks: BackgroundTasks,
        username: str = Depends(authenticate)
    ):
        """Запуск фоновой задачи для лайкания постов ToadBot"""
        if vk_id not in client_manager.clients:
            raise HTTPException(
                status_code=400,
                detail="Аккаунт должен быть запущен (онлайн) для выполнения этой операции."
            )
            
        if vk_id in client_manager.running_like_tasks:
            raise HTTPException(
                status_code=400,
                detail="Задача лайков уже выполняется для этого аккаунта."
            )
            
        background_tasks.add_task(client_manager.like_group_posts, vk_id)
        return {"status": "success", "message": "Задача пролайкивания успешно запущена в фоне."}

    @app.get("/api/accounts/{vk_id}/like_status")
    async def like_status_endpoint(
        vk_id: int,
        username: str = Depends(authenticate)
    ):
        """Получение текущего статуса задачи лайков"""
        status_info = client_manager.like_tasks_status.get(vk_id)
        if not status_info:
            return {
                "is_running": False,
                "stage": "idle",
                "total_posts": 0,
                "collected_posts": 0,
                "to_like_count": 0,
                "liked_count": 0,
                "skipped_count": 0,
                "error": None
            }
        return status_info

    return app
